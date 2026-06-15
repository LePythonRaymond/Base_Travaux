"""DPGF reverse-ingestion parser.

Reads a filled Merci Raymond DPGF (.xlsx) and yields structured line items
ready for matching back to existing products.

Schema assumptions (the master DPGF template, v2):
  - Working zone starts at column AA (client mirror), AD onwards is the
    Merci Raymond cascade (Famille / Sous-cat / Cond / Produit).
  - AG holds the picker string Vincent selected ("Famille — Sous-cat —
    Reference — Conditionnement").
  - AC holds the (mirrored) quantité.
  - BC holds the final PU client (after coefficients).
  - The "DPGF" tab is the operative one.

We try to be robust to small layout variations: if `Paramètres` exposes
Col_Designation / Col_Quantite mappings, we use those for the
client-designation reading.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any, Iterator

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore
    column_index_from_string = None  # type: ignore


# Default column positions on the DPGF tab (1-indexed).
COL_AC = 29  # client-mirror Quantité
COL_AD = 30  # Famille
COL_AE = 31  # Sous-catégorie
COL_AF = 32  # Conditionnement
COL_AG = 33  # Produit (picker string)
COL_AI = 35  # Fournisseur (lookup formula)
COL_AQ = 43  # Fourniture / U
COL_BC = 55  # PU client (per unit)
COL_BE = 57  # Hidden product DB id (NEW; resolved from the picker via lookup)

# Per-line cost+margin breakdown columns (the AK..BC block). Read into
# DpgfLine.breakdown so the product card can show "the coefficients that
# produced this client price" and the project stats can be computed.
BREAKDOWN_COLS = {
    "h_appro_u":        37,  # AK
    "h_pose_u":         38,  # AL
    "nb_uth":           39,  # AM
    "securite_h":       40,  # AN
    "cout_mo_u":        41,  # AO  Coût main-d'œuvre / U
    "cout_humain_tot":  42,  # AP
    "fourniture_u":     43,  # AQ
    "fourniture_tot":   44,  # AR
    "cout_u_fp":        45,  # AS  Coût U fourni posé
    "cout_tot_fp":      46,  # AT  Coût total fourni posé
    "location":         47,  # AU
    "livraison":        48,  # AV
    "install_chantier": 49,  # AW
    "log_gestion":      50,  # AX
    "marge_loc_livr":   51,  # AY
    "marge_humain":     52,  # AZ
    "marge_fourniture": 53,  # BA
    "prix_client_tot":  54,  # BB  Prix client total (PU client × qté)
    "pu_client":        55,  # BC
}

DPGF_SHEET_CANDIDATES = ["DPGF", "DPGF Master", "DPGF Template"]
# The coefficients + rentability recap live on the project-settings tab. It was
# renamed "Paramètres" → "Pilotage de rentabilité"; keep the old names so files
# built before the rename still parse. New names go first so they win.
PARAMS_SHEET_CANDIDATES = [
    "Pilotage de rentabilité", "Pilotage de rentabilite",
    "Paramètres", "Parametres", "Settings",
]

# Paramètres named cells holding the project coefficient snapshot.
PROJECT_COEF_KEYS = [
    "Taux_horaire", "Securite_humain", "Install_chantier", "Log_gestion",
    "Loc_livr_marge", "Humain_marge", "Fourn_marge",
]

# Rentability recap — the AUTHORITATIVE figures the sheet computes via formulas
# on the "Pilotage de rentabilité" tab. Read by identifier (col A) → value
# (col B), exactly like the coefficients. The app no longer re-derives these;
# it reads them (the sheet is ground truth). `marge_pct` cells store the ×100
# value (e.g. 17.75), matching the app's computed marge_pct.
RECAP_GLOBAL_KEYS = {
    "Rent_prix_vente":   "prix_vente",
    "Rent_prix_revient": "prix_revient",
    "Rent_marge_eur":    "marge_eur",
    "Rent_marge_pct":    "marge_pct",
    "Rent_kv":           "kv",
}
RECAP_HORS_SST_KEYS = {
    "Rent_hs_prix_vente":   "prix_vente",
    "Rent_hs_prix_revient": "prix_revient",
    "Rent_hs_marge_eur":    "marge_eur",
    "Rent_hs_marge_pct":    "marge_pct",
    "Rent_hs_kv":           "kv",
}
RECAP_PLANNING_KEYS = {
    "Tps_chantier": "tps_chantier",
    "Personnes":    "personnes",
    "Jours":        "jours",
    "Semaines":     "semaines",
    "Mois":         "mois",
}


class DpgfFormatError(ValueError):
    """Raised when the uploaded workbook doesn't match the Merci Raymond
    DPGF template (no DPGF tab, wrong column layout, etc.). The caller
    surfaces the message to the user — these are expected user errors,
    not crashes."""


@dataclass
class DpgfLine:
    """One parseable line of a filled DPGF."""

    row_index: int                       # 1-indexed sheet row
    client_designation: str = ""         # text mirrored from the client zone
    unit: str | None = None              # mirrored client unit
    quantity: float | None = None        # mirrored Col_Quantite cell
    picker: str | None = None            # AG selected picker string
    famille: str | None = None           # parsed from picker
    sous_cat: str | None = None
    reference_name: str | None = None
    conditionnement: str | None = None
    # `pu_client` is the unit price accepted by the client — read from
    # col BC of the signed DPGF that's being re-ingested. Whether the
    # client renegotiated it during the quote phase is not something this
    # ingestion needs to detect: the workbook we receive is already the
    # accepted one.
    pu_client: float | None = None       # BC value (accepted PU)
    pu_fourniture: float | None = None   # AQ value (per-unit fourniture cost the row used)
    fournisseur: str | None = None       # AI value if formula returned one
    # Stable product DB id resolved from the picker via a hidden lookup column
    # (BE). None for older DPGFs without the column → caller falls back to
    # picker/fuzzy matching.
    product_id: int | None = None
    # Full per-line cost+margin breakdown (the AK..BC columns). Empty for
    # lines that carry no cost chain.
    breakdown: dict[str, float] = field(default_factory=dict)
    raw: dict[str, Any] = field(default_factory=dict)

    @property
    def client_price_differs_from_supplier(self) -> bool:
        """True when the client PU (BC) is meaningfully different from the
        supplier cost (AQ) on the same row. When they're equal, logging
        the BC value separately would just duplicate the cost update — so
        the page skips the validation point in that case. 1-cent
        tolerance to ignore spreadsheet rounding noise."""
        if self.pu_client is None or self.pu_fourniture is None:
            return False
        return abs(self.pu_client - self.pu_fourniture) > 0.01


def parse_dpgf(xlsx_bytes: bytes) -> list[DpgfLine]:
    """Parse a filled DPGF xlsx. Returns rows with at least quantity OR
    picker OR client_designation populated. Empty header / footer rows are
    skipped.

    `data_only=True` so we read FORMULA RESULTS, not the formulas themselves
    (the xlsx must have been opened in Excel/Sheets once for cached values
    to exist — Google Sheets stores them on download).
    """
    if load_workbook is None:
        raise RuntimeError("openpyxl is required for DPGF parsing")

    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
    sheet = None
    for name in DPGF_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        # No "DPGF" tab present → this is almost certainly not a Merci
        # Raymond DPGF workbook. Refuse rather than silently parse the
        # wrong sheet (which would produce noise but no error).
        raise DpgfFormatError(
            "Onglet « DPGF » introuvable dans le classeur. "
            "Le fichier doit être un classeur Merci Raymond DPGF contenant "
            "un onglet nommé « DPGF » (ou « DPGF Master » / « DPGF Template »). "
            f"Onglets trouvés : {', '.join(wb.sheetnames) or '(aucun)'}."
        )

    # Try to read the client-zone column mapping from Paramètres if available.
    col_desig = _resolve_client_column(wb, "Col_Designation", default_col=2)  # B
    col_unite = _resolve_client_column(wb, "Col_Unite", default_col=3)         # C
    col_qte = _resolve_client_column(wb, "Col_Quantite", default_col=5)        # E

    lines: list[DpgfLine] = []
    # Data rows are typically 3..502 per template.
    for r in range(3, min(sheet.max_row + 1, 600)):
        qty_val = sheet.cell(row=r, column=COL_AC).value
        picker_val = sheet.cell(row=r, column=COL_AG).value
        # Read client-zone via mapped columns, then fall back to AA-AB if blank
        client_desig = (
            sheet.cell(row=r, column=col_desig).value
            or sheet.cell(row=r, column=27).value   # AA fallback
        )
        unit_val = (
            sheet.cell(row=r, column=col_unite).value
            or sheet.cell(row=r, column=28).value   # AB fallback
        )
        client_qty = sheet.cell(row=r, column=col_qte).value

        # Use client_qty if AC mirror was empty (happens when formula failed)
        if qty_val is None and isinstance(client_qty, (int, float)):
            qty_val = client_qty

        # Skip rows with nothing to ingest
        if not qty_val and not picker_val and not client_desig:
            continue
        # Skip rows whose qty is a string (header/section row)
        if qty_val is not None and not isinstance(qty_val, (int, float)):
            qty_val = None

        line = DpgfLine(
            row_index=r,
            client_designation=str(client_desig or "").strip(),
            unit=(str(unit_val).strip() if unit_val else None),
            quantity=(float(qty_val) if qty_val is not None else None),
        )

        if picker_val and isinstance(picker_val, str) and picker_val.strip():
            line.picker = picker_val.strip()
            parts = [p.strip() for p in line.picker.split("—")]
            if len(parts) == 4:
                line.famille, line.sous_cat, line.reference_name, line.conditionnement = parts
            else:
                # Try old 3-part picker as a fallback
                if len(parts) == 3:
                    line.famille, line.reference_name, line.conditionnement = parts

        pu_client = sheet.cell(row=r, column=COL_BC).value
        if isinstance(pu_client, (int, float)):
            line.pu_client = float(pu_client)

        pu_fourn = sheet.cell(row=r, column=COL_AQ).value
        if isinstance(pu_fourn, (int, float)):
            line.pu_fourniture = float(pu_fourn)

        fournisseur = sheet.cell(row=r, column=COL_AI).value
        if isinstance(fournisseur, str) and fournisseur.strip():
            line.fournisseur = fournisseur.strip()

        # Hidden product DB id (BE). Present only on DPGFs built from the
        # current template; older files leave it None → fuzzy match.
        pid_val = sheet.cell(row=r, column=COL_BE).value
        if isinstance(pid_val, (int, float)) and float(pid_val) > 0:
            line.product_id = int(pid_val)
        elif isinstance(pid_val, str) and pid_val.strip().isdigit():
            line.product_id = int(pid_val.strip())

        # Full per-line cost/margin breakdown.
        bd: dict[str, float] = {}
        for key, col in BREAKDOWN_COLS.items():
            v = sheet.cell(row=r, column=col).value
            if isinstance(v, (int, float)):
                bd[key] = float(v)
        line.breakdown = bd

        line.raw = {
            "AC": qty_val,
            "AG": picker_val,
            "AI": fournisseur,
            "AQ": pu_fourn,
            "BC": pu_client,
            "BE": pid_val,
        }
        lines.append(line)

    return lines


def parse_project_meta(xlsx_bytes: bytes) -> dict[str, Any]:
    """Extract project-level metadata from a DPGF workbook:

      • coefficients — the project coefficient snapshot (Taux_horaire, the
        margins…) read from the Paramètres named cells.
      • computed — rentability computed from the line items:
        prix_vente = Σ(PU client × qté), prix_revient = Σ(coût total fourni
        posé + loc + livr + install + log), marge_eur, marge_pct, kv.
      • recap — the AUTHORITATIVE rentability the SHEET computes via formulas
        on the "Pilotage de rentabilité" tab (GLOBAL + Hors-SST + the Tps-
        chantier planning line), read by identifier from fixed cells. The
        sheet is ground truth; `computed` is kept only as a silent cross-check.
        Falls back to the legacy label-search at the bottom of the DPGF tab for
        older workbooks that predate the recap block.

    Never raises on a malformed sheet — returns whatever it can so ingestion
    is never blocked by stats parsing.
    """
    out: dict[str, Any] = {"coefficients": {}, "computed": {}, "recap": {}}
    if load_workbook is None:
        return out
    try:
        wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
    except Exception:  # noqa: BLE001
        return out

    # ── coefficients (Paramètres named cells) ──
    out["coefficients"] = _read_project_coefficients(wb)

    # ── computed totals from the lines ──
    try:
        lines = parse_dpgf(xlsx_bytes)
    except Exception:  # noqa: BLE001
        lines = []
    prix_vente = 0.0
    prix_revient = 0.0
    for ln in lines:
        bd = ln.breakdown or {}
        # Selling side: prefer the per-line "prix client total" (already × qté),
        # else PU client × quantity.
        if bd.get("prix_client_tot"):
            prix_vente += float(bd["prix_client_tot"])
        elif ln.pu_client is not None and ln.quantity is not None:
            prix_vente += float(ln.pu_client) * float(ln.quantity)
        # Cost side: coût total fourni posé + loc + livr + install + log
        # (the full prix-de-revient before margins). Fall back to fourniture
        # × qté when the breakdown is absent.
        cost = (
            bd.get("cout_tot_fp", 0.0)
            + bd.get("location", 0.0)
            + bd.get("livraison", 0.0)
            + bd.get("install_chantier", 0.0)
            + bd.get("log_gestion", 0.0)
        )
        if cost:
            prix_revient += cost
        elif ln.pu_fourniture is not None and ln.quantity is not None:
            prix_revient += float(ln.pu_fourniture) * float(ln.quantity)
    marge = prix_vente - prix_revient
    out["computed"] = {
        "prix_vente": round(prix_vente, 2),
        "prix_revient": round(prix_revient, 2),
        "marge_eur": round(marge, 2),
        "marge_pct": round(marge / prix_vente * 100, 2) if prix_vente else None,
        "kv": round(prix_vente / prix_revient, 4) if prix_revient else None,
        "n_lines": len(lines),
    }

    # ── recap block — AUTHORITATIVE (sheet formulas = ground truth) ──
    # 1. fixed identifier→value cells on the "Pilotage de rentabilité" tab
    #    (what the sheet now exposes); 2. fallback: legacy label-search at the
    #    bottom of the DPGF tab, for workbooks built before the recap block.
    recap = _read_project_recap_cells(wb)
    if not recap:
        sheet = None
        for name in DPGF_SHEET_CANDIDATES:
            if name in wb.sheetnames:
                sheet = wb[name]
                break
        if sheet is not None:
            recap = _capture_recap(sheet)
    out["recap"] = recap

    return out


def _read_project_recap_cells(wb) -> dict[str, Any]:
    """Read the authoritative rentability recap from the "Pilotage de
    rentabilité" tab: col A = identifier, col B = value (same convention as the
    coefficients). The sheet computes everything — incl. the Hors-SST split via
    the SST tick-box on the DPGF tab — so we only READ. Returns {} when the tab
    or the identifiers are absent, so the caller can fall back to label-search.

    Output shape (keys present only when found):
        {prix_vente, prix_revient, marge_eur, marge_pct, kv,
         hors_sst: {prix_vente, prix_revient, marge_eur, marge_pct, kv},
         tps_chantier, personnes, jours, semaines, mois}
    """
    params = None
    for name in PARAMS_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            params = wb[name]
            break
    if params is None:
        return {}

    # Layout on the recap rows: col A = human label, col B = value, col C =
    # machine identifier (the column the sheet hides). We match on the col-C
    # identifier and read the col-B value. (Coefficients above keep A=id/B=val,
    # read separately by _read_project_coefficients — no collision.)
    out: dict[str, Any] = {}
    hors: dict[str, float] = {}
    for row in range(1, min(params.max_row + 1, 80)):
        ident = params.cell(row=row, column=3).value
        if not ident:
            continue
        key = str(ident).strip()
        val = _to_number(params.cell(row=row, column=2).value)
        if val is None:
            continue
        if key in RECAP_GLOBAL_KEYS:
            out[RECAP_GLOBAL_KEYS[key]] = val
        elif key in RECAP_HORS_SST_KEYS:
            hors[RECAP_HORS_SST_KEYS[key]] = val
        elif key in RECAP_PLANNING_KEYS:
            out[RECAP_PLANNING_KEYS[key]] = val

    if hors:
        out["hors_sst"] = hors
    return out


def _read_project_coefficients(wb) -> dict[str, float]:
    """Read the project coefficient snapshot from the Paramètres tab
    (Identifiant in col A, Valeur in col B)."""
    params = None
    for name in PARAMS_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            params = wb[name]
            break
    if params is None:
        return {}
    coefs: dict[str, float] = {}
    for row in range(1, min(params.max_row + 1, 40)):
        ident = params.cell(row=row, column=1).value
        if ident and str(ident).strip() in PROJECT_COEF_KEYS:
            val = params.cell(row=row, column=2).value
            if isinstance(val, (int, float)):
                coefs[str(ident).strip()] = float(val)
    return coefs


def _to_number(v: Any) -> float | None:
    """Parse a cell value that may be a number or a French-formatted string
    ('1 247 000,00 €', '17,81%')."""
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str):
        return None
    s = v.replace("€", "").replace("%", "")
    s = re.sub(r"\s", "", s).replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _capture_recap(sheet) -> dict[str, Any]:
    """Scan the bottom rows of the DPGF tab for the rentability recap labels
    and grab the numeric value in the next non-empty cell to the right.
    Best-effort: returns whatever it finds, splitting GLOBAL vs Hors-SST."""
    labels = {
        "prix vente": "prix_vente",
        "prix de revient": "prix_revient",
        "marge": "marge_eur",
        "kv": "kv",
    }
    last = min(sheet.max_row, 700)
    start = max(3, last - 80)  # only the bottom band holds the recap
    block_global: dict[str, float] = {}
    block_hors: dict[str, float] = {}
    in_hors = False
    for r in range(start, last + 1):
        for c in range(1, min(sheet.max_column + 1, 60)):
            cell = sheet.cell(row=r, column=c).value
            if not isinstance(cell, str):
                continue
            txt = cell.strip().lower()
            if "hors sst" in txt:
                in_hors = True
            key = None
            for lbl, k in labels.items():
                if txt == lbl or txt.startswith(lbl):
                    key = k
                    break
            if key is None:
                continue
            # value = first numeric cell to the right on the same row
            val = None
            for cc in range(c + 1, min(sheet.max_column + 1, 60)):
                val = _to_number(sheet.cell(row=r, column=cc).value)
                if val is not None:
                    break
            if val is None:
                continue
            (block_hors if in_hors else block_global)[key] = val
    recap: dict[str, Any] = dict(block_global)
    if block_hors:
        recap["hors_sst"] = block_hors
    return recap


def _resolve_client_column(wb, named_range: str, default_col: int) -> int:
    """Look up a Paramètres-cell name (Col_Designation etc.) and return its
    1-indexed column position. Returns `default_col` if not found.
    """
    if column_index_from_string is None:
        return default_col

    # Find Paramètres sheet
    params_sheet = None
    for name in PARAMS_SHEET_CANDIDATES:
        if name in wb.sheetnames:
            params_sheet = wb[name]
            break
    if params_sheet is None:
        return default_col

    # The template stores Identifiant in column A, Valeur in column B
    # (rows 16-18 hold Col_Designation, Col_Unite, Col_Quantite).
    for row in range(15, 25):
        identifier = params_sheet.cell(row=row, column=1).value
        if identifier and str(identifier).strip() == named_range:
            value = params_sheet.cell(row=row, column=2).value
            if value and isinstance(value, str):
                try:
                    return column_index_from_string(value.strip().upper())
                except Exception:  # noqa: BLE001
                    return default_col
            break
    return default_col


def stats(lines: list[DpgfLine]) -> dict[str, int]:
    """Return rough counts: total / with_picker / with_qty / matched / unmatched."""
    total = len(lines)
    with_picker = sum(1 for l in lines if l.picker)
    with_qty = sum(1 for l in lines if l.quantity is not None)
    return {"total": total, "with_picker": with_picker, "with_qty": with_qty}
