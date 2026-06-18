"""Load a workbook (xlsx or csv export) into uniform in-memory Grids.

A Grid is a thin 1-indexed view over a 2D value array so every downstream module
(anchors, multiplier, rows) works identically on .xlsx sheets and .csv exports.
openpyxl is opened with data_only=True so we read cached formula *values*; cells
Excel stored as dates stay as datetime objects (needed for date-corruption flags).
"""

from __future__ import annotations

import csv
import warnings
from dataclasses import dataclass, field

import openpyxl
from openpyxl.utils import get_column_letter

from .config import normalize

warnings.filterwarnings("ignore", message="Cannot parse header or footer")


@dataclass
class Grid:
    name: str
    state: str = "visible"
    rows: list[list[object]] = field(default_factory=list)  # 0-indexed internally
    merged: list[tuple[int, int, int, int]] = field(default_factory=list)  # 1-indexed bbox

    @property
    def nrows(self) -> int:
        return len(self.rows)

    @property
    def ncols(self) -> int:
        return max((len(r) for r in self.rows), default=0)

    def cell(self, r: int, c: int) -> object:
        """1-indexed value access; out-of-range → None."""
        if 1 <= r <= len(self.rows):
            row = self.rows[r - 1]
            if 1 <= c <= len(row):
                return row[c - 1]
        return None

    def norm(self, r: int, c: int) -> str:
        return normalize(self.cell(r, c))

    def row_norms(self, r: int) -> list[str]:
        return [self.norm(r, c) for c in range(1, self.ncols + 1)]

    @staticmethod
    def col_letter(c: int) -> str:
        return get_column_letter(c)


def _grid_from_ws(ws) -> Grid:
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    merged = []
    for mr in ws.merged_cells.ranges:
        merged.append((mr.min_row, mr.min_col, mr.max_row, mr.max_col))
    state = getattr(ws, "sheet_state", "visible")
    return Grid(name=ws.title, state=state, rows=rows, merged=merged)


def load_grids(path: str) -> dict[str, Grid]:
    """Return {sheet_name: Grid}. CSV files yield a single grid named after the file."""
    if path.lower().endswith(".csv"):
        with open(path, newline="", encoding="utf-8", errors="replace") as fh:
            rows = [list(r) for r in csv.reader(fh)]
        name = path.rsplit("/", 1)[-1]
        return {name: Grid(name=name, rows=rows)}

    wb = openpyxl.load_workbook(path, data_only=True)
    grids: dict[str, Grid] = {}
    for ws in wb.worksheets:
        grids[ws.title] = _grid_from_ws(ws)
    wb.close()
    return grids


def expand_merges(grid: Grid) -> dict[tuple[int, int], tuple[int, int]]:
    """Map every covered (row, col) to the merge's top-left (where the value lives).

    Lets us read a band label (stored only in the merge's top-left) as covering all
    its sub-columns, so 'FOURNITURE' aligns over the 'Fourniture/U' sub-header.
    """
    cover: dict[tuple[int, int], tuple[int, int]] = {}
    for (r0, c0, r1, c1) in grid.merged:
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                cover[(r, c)] = (r0, c0)
    return cover
