"""Load the APPROVED review workbook into the DB — idempotent, FK-ordered.

Reads only rows whose `approve` is ticked. Runs inside a single
``lib.db.transaction(ingestion_source='historical_dpgf')`` so the
``log_price_change`` trigger attributes every auto-logged supplier-cost row
correctly. Re-running over the same approved workbook is a no-op (every step is
ON CONFLICT / existence-guarded).

    python -m dpgf_corpus_etl.run_load --review review.xlsx --dry-run
    python -m dpgf_corpus_etl.run_load --review review.xlsx

(Kept out of run.py so `extract` never imports Streamlit/DB. Invoke via run_load.py.)
"""

from __future__ import annotations

import json
import os
import sys

import openpyxl

_APPROVE_TRUE = {"✓", "x", "X", "oui", "yes", "y", "1", "true", "vrai", "ok"}


def _approved(v) -> bool:
    return v is not None and str(v).strip() in _APPROVE_TRUE


def _rows(ws):
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield dict(zip(headers, row))


def load_approved(review_path: str, *, dry_run: bool = False, actor: str = "corpus_etl") -> dict:
    # Import here so `extract` never needs Streamlit/DB on the path.
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "..", "streamlit_app"))

    wb = openpyxl.load_workbook(review_path, data_only=True)
    all_products = list(_rows(wb["products"]))
    products = [r for r in all_products if _approved(r.get("approve"))]
    # NOT auto-approved → staged in ingestion_queue as 'needs_info' (À vérifier),
    # so the team verifies the unit / forfait / junk in the app before it goes live.
    flagged = [r for r in all_products if not _approved(r.get("approve")) and r.get("reference_name")]
    labor = [r for r in _rows(wb["labor_norms"]) if _approved(r.get("approve"))]
    taxonomy = [r for r in _rows(wb["taxonomy"]) if _approved(r.get("approve"))]
    # all harvested suppliers are real curated names → load them all (handy even
    # when not yet attached to a product); plus any referenced by products.
    suppliers = {r.get("supplier_name") for r in _rows(wb["suppliers"]) if r.get("supplier_name")}
    suppliers |= {p.get("supplier_name") for p in products + flagged if p.get("supplier_name")}
    suppliers.discard(None)
    suppliers.add("Fournisseur inconnu")

    report = {"families": 0, "suppliers": 0, "labor_norms": 0, "taxonomy": 0,
              "products_new": 0, "products_updated": 0, "price_history": 0,
              "queued_a_verifier": 0, "skipped": 0, "dry_run": dry_run}

    if dry_run:
        report["families"] = len({p.get("family") for p in products if p.get("family")})
        report["suppliers"] = len(suppliers)
        report["labor_norms"] = len(labor)
        report["taxonomy"] = len(taxonomy)
        report["products_new"] = len(products)  # upper bound
        report["price_history"] = sum(1 for _ in _rows(wb["price_history"]))
        report["queued_a_verifier"] = len(flagged)
        return report

    from sqlalchemy import text  # noqa: E402  (imported here so --dry-run needs no DB/Streamlit)
    from lib.db import transaction  # noqa: E402
    with transaction(ingestion_source="historical_dpgf", ingestion_actor=actor) as conn:
        # 1) families
        fam_id: dict[str, int] = {}
        for name in sorted({p.get("family") for p in products if p.get("family")}):
            rid = conn.execute(text(
                "INSERT INTO product_families (name) VALUES (:n) "
                "ON CONFLICT (name) DO UPDATE SET name = EXCLUDED.name RETURNING id"),
                {"n": name}).scalar()
            fam_id[name] = int(rid)
            report["families"] += 1

        # 2) suppliers
        sup_id: dict[str, int] = {}
        for name in sorted(suppliers):
            rid = conn.execute(text(
                "INSERT INTO suppliers (name) VALUES (:n) "
                "ON CONFLICT (name) DO UPDATE SET name = EXCLUDED.name RETURNING id"),
                {"n": name}).scalar()
            sup_id[name] = int(rid)
            report["suppliers"] += 1

        # 3) labor_norms (full computed tiers) — DO NOTHING keeps hand-tuned norms
        norm_id: dict[str, int] = {}
        for n in labor:
            rid = conn.execute(text(
                """
                INSERT INTO labor_norms
                  (task_name, unit_type, nombre_uth_default, heure_u_pose_default,
                   tier_1_label, tier_1_heure_u_decharge,
                   tier_2_label, tier_2_heure_u_decharge,
                   tier_3_label, tier_3_heure_u_decharge, notes)
                VALUES (:t, :u, :uth, :pose, 'facile', :t1, 'moyen', :t2, 'difficile', :t3,
                        :notes)
                ON CONFLICT (task_name) DO UPDATE SET task_name = EXCLUDED.task_name
                RETURNING id
                """),
                {"t": n["task_name"], "u": n.get("unit_type") or "u",
                 "uth": n.get("nombre_uth_default") or 1, "pose": n.get("heure_u_pose_default") or 0,
                 "t1": n.get("tier_1_facile") or 0, "t2": n.get("tier_2_moyen") or 0,
                 "t3": n.get("tier_3_difficile") or 0,
                 "notes": f"historical_dpgf n_obs={n.get('n_obs')} {n.get('flags') or ''}".strip()}).scalar()
            norm_id[n["task_name"]] = int(rid)
            report["labor_norms"] += 1
        default_norm = conn.execute(text(
            "SELECT id FROM labor_norms WHERE task_name = 'Norme par défaut (à classifier)'")).scalar()

        # 4) taxonomy triplets
        for t in taxonomy:
            fid = fam_id.get(t.get("family"))
            if not fid:
                continue
            conn.execute(text(
                "INSERT INTO product_taxonomy (family_id, subcategory, packaging, created_by, notes) "
                "VALUES (:f,:s,:p,'historical_dpgf','corpus load') "
                "ON CONFLICT (family_id, subcategory, packaging) DO NOTHING"),
                {"f": fid, "s": t.get("subcategory") or "À classifier", "p": t.get("packaging") or "Standard"})
            report["taxonomy"] += 1

        # 5) + 6) products + price_history
        for p in products:
            fam = p.get("family")
            if not fam or fam not in fam_id:
                report["skipped"] += 1
                continue
            fid = fam_id[fam]
            sub = (p.get("subcategory") or "À classifier").strip()
            pkg = (p.get("packaging") or "Standard").strip() or "Standard"
            unit = (p.get("unit_type") or "u").strip() or "u"
            supplier = p.get("supplier_name") or "Fournisseur inconnu"
            sid = sup_id.get(supplier) or sup_id["Fournisseur inconnu"]
            lid = norm_id.get(p.get("labor_task")) or default_norm
            cost = p.get("cost_ht")
            if cost is None:
                report["skipped"] += 1
                continue

            conn.execute(text(
                "INSERT INTO product_taxonomy (family_id, subcategory, packaging, created_by, notes) "
                "VALUES (:f,:s,:p,'historical_dpgf','corpus load') "
                "ON CONFLICT (family_id, subcategory, packaging) DO NOTHING"),
                {"f": fid, "s": sub, "p": pkg})

            # per-product source reference drives the trigger-logged history row
            conn.execute(text("SELECT set_config('app.ingestion_reference', :r, true)"),
                         {"r": (p.get("source_files") or "historical_dpgf")[:200]})

            attrs = p.get("attributes_json") or "{}"
            try:
                json.loads(attrs)
            except Exception:
                attrs = "{}"
            row = conn.execute(text(
                """
                INSERT INTO products
                  (reference_name, family_id, subcategory, packaging, unit_type,
                   supplier_id, labor_norm_id, brand, material, attributes, cost_ht, notes)
                VALUES
                  (:ref, :fid, :sub, :pkg, :unit, :sid, :lid, :brand, :material,
                   CAST(:attrs AS jsonb), :cost, :notes)
                ON CONFLICT (reference_name, packaging, supplier_id) DO UPDATE
                  SET cost_ht = EXCLUDED.cost_ht, last_price_update = now()
                RETURNING id, (xmax = 0) AS inserted
                """),
                {"ref": p["reference_name"], "fid": fid, "sub": sub, "pkg": pkg, "unit": unit,
                 "sid": sid, "lid": lid, "brand": p.get("brand") or None,
                 "material": p.get("material") or None, "attrs": attrs, "cost": cost,
                 "notes": f"historical_dpgf: {p.get('source_files') or ''}"[:500]}).mappings().first()
            pid, inserted = int(row["id"]), bool(row["inserted"])
            report["products_new" if inserted else "products_updated"] += 1

            # initial history row for a brand-new product (trigger only fires on UPDATE)
            if inserted:
                conn.execute(text(
                    "INSERT INTO price_history (product_id, cost_ht, source, source_reference, recorded_by) "
                    "VALUES (:pid,:cost,'historical_dpgf',:ref,:by)"),
                    {"pid": pid, "cost": cost, "ref": (p.get("source_files") or "")[:200], "by": actor})
                report["price_history"] += 1

        # 7) flagged products → ingestion_queue ('needs_info' = À vérifier).
        #    They surface in the À classifier page → "Ingestion en attente" tab,
        #    pre-filled, for the team to set the unit / confirm the forfait / reject.
        #    Re-running replaces the prior historical_dpgf pending set (idempotent),
        #    without touching rows a human already approved/rejected.
        conn.execute(text(
            "DELETE FROM ingestion_queue WHERE source='historical_dpgf' "
            "AND status IN ('pending','needs_info')"))
        for p in flagged:
            supplier = p.get("supplier_name") or "Fournisseur inconnu"
            sid = sup_id.get(supplier)
            lid = norm_id.get(p.get("labor_task"))
            payload = json.dumps(
                {k: (str(v) if v is not None else None) for k, v in p.items()},
                ensure_ascii=False)
            conn.execute(text(
                """
                INSERT INTO ingestion_queue
                  (source, source_reference, raw_payload, candidate_reference_name,
                   candidate_family_hint, candidate_packaging, candidate_unit_type,
                   candidate_supplier_id, candidate_supplier_hint,
                   candidate_labor_norm_id, candidate_labor_hint, candidate_cost_ht,
                   status, review_notes)
                VALUES
                  ('historical_dpgf', :ref0, CAST(:payload AS jsonb), :ref, :fam, :pkg, :unit,
                   :sid, :suph, :lid, :lhint, :cost, 'needs_info', :note)
                """),
                {"ref0": (p.get("source_files") or "")[:200], "payload": payload,
                 "ref": p.get("reference_name"), "fam": p.get("family") or None,
                 "pkg": p.get("packaging") or None, "unit": p.get("unit_type") or None,
                 "sid": sid, "suph": None if sid else supplier,
                 "lid": lid, "lhint": None if lid else p.get("labor_task"),
                 "cost": p.get("cost_ht"), "note": _flag_note(p.get("flags"), p.get("unit_type"))})
            report["queued_a_verifier"] += 1

    return report


def _flag_note(flags: str | None, unit) -> str:
    """Human-readable 'À vérifier' note from the extractor flags."""
    f = flags or ""
    parts = ["À vérifier — chargement DPGF historique."]
    if "lump_unit" in f:
        parts.append("Prix au forfait (Ft) : confirmer que c'est un prix global, pas unitaire.")
    if "unit_unmapped" in f:
        parts.append(f"Unité non reconnue ({unit or '?'}) : préciser l'unité.")
    if "cost_date_corruption" in f:
        parts.append("Cellule de prix corrompue en date : ressaisir le coût.")
    if not f:
        parts.append("Confiance faible / possible en-tête de section : vérifier que c'est bien un produit.")
    return " ".join(parts)
