"""Write the human-review workbook Vincent/Taddeo approve before any DB load.

Tabs mirror the target tables (suppliers, labor_norms, products, price_history,
taxonomy) plus a _qa tab. The loader reads back only rows whose `approve` = ✓.
"""

from __future__ import annotations

import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .config import UNIT_TYPES

_HEAD = Font(bold=True, color="FFFFFF")
_FILL = PatternFill("solid", fgColor="2E5A34")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _sheet(wb: Workbook, title: str, headers: list[str]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = _HEAD
        cell.fill = _FILL
    ws.freeze_panes = "A2"
    return ws


def write_review(path: str, *, products, labor_norms, suppliers, taxonomy, qa,
                 price_history) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    ws = _sheet(wb, "suppliers", ["approve", "supplier_name", "source_files", "reviewer_notes"])
    for s in sorted(suppliers):
        ws.append(["", s, "", ""])

    ws = _sheet(wb, "labor_norms", [
        "approve", "labor_id", "task_name", "unit_type", "nombre_uth_default", "heure_u_pose_default",
        "tier_1_facile", "tier_2_moyen", "tier_3_difficile", "n_obs",
        "decharge_min", "decharge_median", "decharge_max", "flags", "reviewer_notes"])
    for n in sorted(labor_norms, key=lambda x: -(x.get("n_obs") or 0)):
        # sure norms are pre-approved (✓); reviewer can still untick
        ws.append(["✓", n.get("labor_id", ""), n["task_name"], n["unit_type"], n["nombre_uth_default"],
                   n["heure_u_pose_default"], n["tier_1_heure_u_decharge"],
                   n["tier_2_heure_u_decharge"], n["tier_3_heure_u_decharge"], n["n_obs"],
                   n["decharge_min"], n["decharge_median"], n["decharge_max"], n["flags"], ""])

    ws = _sheet(wb, "products", [
        "approve", "reference_name", "family", "subcategory", "packaging", "unit_type",
        "cost_ht", "brand", "material", "attributes_json", "supplier_name", "labor_task",
        "n_sources", "cost_min", "cost_max", "confidence", "flags", "is_option",
        "source_files", "source_cells", "reviewer_notes"])
    for p in sorted(products, key=lambda x: (x.family or "zz", x.reference_name)):
        costs = [float(c) for c in p.costs] if p.costs else []
        ws.append([
            "✓" if (p.confidence >= 0.6 and not p.flags and p.cost_ht is not None) else "",
            p.reference_name, p.family or "", p.subcategory or "", p.packaging or "",
            p.unit or "", float(p.cost_ht) if p.cost_ht is not None else None,
            p.brand or "", p.material or "",
            json.dumps(p.attributes, ensure_ascii=False) if p.attributes else "",
            p.supplier or "", p.labor_task or "", p.n_sources,
            min(costs) if costs else None, max(costs) if costs else None,
            round(p.confidence, 2), ",".join(p.flags), "yes" if p.is_option else "",
            "; ".join(sorted(p.source_files)), " | ".join(p.source_cells[:6]), ""])
    # unit dropdown
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(UNIT_TYPES), allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"F2:F{ws.max_row}")

    ws = _sheet(wb, "price_history", [
        "product_ref", "unit_type", "cost_ht", "source", "source_reference", "flags"])
    for h in price_history:
        ws.append([h["ref"], h["unit"], h["cost"], "historical_dpgf", h["file"], h.get("flags", "")])

    ws = _sheet(wb, "taxonomy", ["approve", "family", "subcategory", "packaging", "from_products", "reviewer_notes"])
    for (fam, sub, pkg), cnt in sorted(taxonomy.items()):
        ws.append(["", fam, sub, pkg, cnt, ""])

    ws = _sheet(wb, "_qa", [
        "file", "priced_sheets", "cost_cols", "confidence", "multiplier_rows",
        "n_lines", "n_with_cost", "n_with_labor", "warnings"])
    for q in qa:
        ws.append([q["file"], q["priced_sheets"], q["cost_cols"], q["confidence"],
                   q["multiplier_rows"], q["n_lines"], q["n_cost"], q["n_labor"], q["warnings"]])

    # column widths (rough)
    for ws in wb.worksheets:
        for col in ws.columns:
            width = min(46, max(10, max((len(str(c.value)) if c.value else 0) for c in col) + 2))
            ws.column_dimensions[col[0].column_letter].width = width
            for c in col:
                c.alignment = _WRAP
    wb.save(path)
